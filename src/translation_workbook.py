"""New-workbook-only translation review workbooks from completed evidence.

The writer consumes immutable job, layout, row, and result evidence.  It never
selects OCR text, invokes a provider, or updates a reviewed workbook.
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from excel_exporter import (
    AI_TRANSLATION_FILL, BASE_ROW_HEIGHT, HEADER_FILL, HEADER_ROW_HEIGHT,
    MODIFIED_TRANSLATION_FILL,
    SOURCE_REFERENCE_FILL, TABLE_HEADER_ROW, TABLE_START_ROW,
    TEXT_COLUMN_WIDTH, TITLE_TEXT, VERIFIED_COLUMN_WIDTH, VERIFIED_FILL,
    _estimate_row_height,
)
from translation_contract import TranslationResult, TranslationStatus
from translation_job import TranslationJob
from translation_output_plan import PlannedSheet, TranslationOutputLayout
from translation_review import (HumanTranslationReview, TranslationReviewAssessment, assess_translation_results,
    human_review_status_display, resolve_reviewed_translation, review_status_display)


_METADATA_SHEET_NAME = "_VideoText_Metadata"
_SCHEMA_VERSION = "2"
_TABLE_HEADERS = (
    "Slide", "Source OCR Text", "Original AI Translation", "Verified Translation",
    "Human Review Status", "Translation Review Status", "Review Reasons",
    "Reviewer Notes", "Target Language", "Provider", "Model",
)
_REVIEW_GUIDANCE = ("AI-generated translations require human review. Items marked Review Recommended "
                    "have additional automated warning signals and may warrant particular attention. "
                    "Use Verified Translation and Human Review Status for human review; source OCR and "
                    "the original AI translation remain unchanged.")
_VERIFIED_TRANSLATION_NOTE = (
    "Enter the human-verified translation here. The original AI translation remains unchanged."
)
_SENSITIVE_METADATA_KEY = re.compile(r"(?:api.?key|token|secret|password|credential|authorization)", re.IGNORECASE)
_SENSITIVE_ERROR = re.compile(r"(?:bearer\s+|sk-[A-Za-z0-9_-]+|(?:api[_ -]?key|token|secret|password)\s*[=:]\s*)[^\s,;]+", re.IGNORECASE)


def _required(value: str, field_name: str) -> None:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{field_name} is required.")


@dataclass(frozen=True)
class TranslationWorkbookRow:
    """Immutable sheet-row evidence already selected by upstream translation work."""

    source_item_id: str
    source_reference: str
    request_id: str
    target_language: str
    slide_value: int | str
    original_text: str
    ordering_index: int

    def __post_init__(self) -> None:
        for value, name in ((self.source_item_id, "source_item_id"), (self.source_reference, "source_reference"),
                            (self.request_id, "request_id"), (self.target_language, "target_language"),
                            (self.original_text, "original_text")):
            _required(value, name)
        if not isinstance(self.slide_value, (int, str)) or isinstance(self.slide_value, bool):
            raise ValueError("slide_value must be an integer or display string.")
        if not isinstance(self.ordering_index, int) or self.ordering_index < 0:
            raise ValueError("ordering_index must be a non-negative integer.")


@dataclass(frozen=True)
class TranslationWorkbookWriteResult:
    """Immutable summary of successfully created new translation workbooks."""

    workbook_paths: Mapping[str, Path]
    sheet_count: int
    row_count: int
    success_translation_count: int
    failure_translation_count: int

    def __post_init__(self) -> None:
        object.__setattr__(self, "workbook_paths", MappingProxyType(dict(self.workbook_paths)))


def _safe_metadata(value: Any) -> Any:
    """Return JSON-safe metadata while excluding credential-like fields."""

    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (list, tuple)):
        return [_safe_metadata(item) for item in value]
    if isinstance(value, Mapping):
        return {str(key): _safe_metadata(item) for key, item in value.items()
                if not _SENSITIVE_METADATA_KEY.search(str(key))}
    return "<unsupported metadata>"


def _safe_error(error: str | None) -> str | None:
    """Retain a useful error category while preventing common secret leakage."""

    return _SENSITIVE_ERROR.sub("[redacted]", error) if error else None


def _rows_by_source(rows: Iterable[TranslationWorkbookRow], job: TranslationJob) -> dict[str, tuple[TranslationWorkbookRow, ...]]:
    """Validate source identity and preserve the upstream row ordering."""

    valid_sources = {item.source_item_id for item in job.source_items}
    grouped: dict[str, list[TranslationWorkbookRow]] = {item.source_item_id: [] for item in job.source_items}
    seen_requests: set[str] = set()
    for row in rows:
        if not isinstance(row, TranslationWorkbookRow):
            raise ValueError("source_rows must contain TranslationWorkbookRow values.")
        if row.source_item_id not in valid_sources:
            raise ValueError(f"Workbook row has an unknown source item: {row.source_item_id}.")
        if row.target_language not in job.target_languages:
            raise ValueError(f"Workbook row has an unsupported target language: {row.target_language}.")
        if row.request_id in seen_requests:
            raise ValueError(f"Duplicate expected translation request: {row.request_id}.")
        seen_requests.add(row.request_id)
        grouped[row.source_item_id].append(row)
    return {source_id: tuple(sorted(values, key=lambda value: value.ordering_index)) for source_id, values in grouped.items()}


def _result_map(results: Iterable[TranslationResult]) -> dict[str, TranslationResult]:
    """Map immutable results by request identity, rejecting duplicate evidence."""

    mapped: dict[str, TranslationResult] = {}
    for result in results:
        if not isinstance(result, TranslationResult):
            raise ValueError("translation_results must contain TranslationResult values.")
        if result.request_id in mapped:
            raise ValueError(f"Duplicate translation result for request: {result.request_id}.")
        mapped[result.request_id] = result
    return mapped


def _human_review_map(reviews: Iterable[HumanTranslationReview] | None,
                      request_ids: set[str]) -> dict[str, HumanTranslationReview]:
    """Return one explicit human-review record per result, defaulting to Unreviewed."""

    if reviews is None:
        return {request_id: HumanTranslationReview(request_id) for request_id in request_ids}
    mapped: dict[str, HumanTranslationReview] = {}
    for review in reviews:
        if not isinstance(review, HumanTranslationReview):
            raise ValueError("human_reviews must contain HumanTranslationReview values.")
        if review.request_id in mapped:
            raise ValueError(f"Duplicate human review for request: {review.request_id}.")
        mapped[review.request_id] = review
    if set(mapped) != request_ids:
        identifier = sorted((request_ids - set(mapped)) or (set(mapped) - request_ids))[0]
        raise ValueError(f"Human reviews must match translation results exactly: {identifier}.")
    return mapped


def _validate_sheet_evidence(sheet: PlannedSheet, rows: tuple[TranslationWorkbookRow, ...], results: Mapping[str, TranslationResult]) -> tuple[tuple[TranslationWorkbookRow, TranslationResult], ...]:
    """Strictly match a sheet's expected rows to exactly one compatible result."""

    matched: list[tuple[TranslationWorkbookRow, TranslationResult]] = []
    for row in rows:
        if row.target_language != sheet.target_language:
            continue
        try:
            result = results[row.request_id]
        except KeyError as error:
            raise ValueError(f"Missing translation result for request: {row.request_id}.") from error
        request = result.request
        if request.target_language != sheet.target_language:
            raise ValueError(f"Translation result target language does not match sheet: {row.request_id}.")
        if request.provenance.source_reference != row.source_reference:
            raise ValueError(f"Translation result source reference does not match row: {row.request_id}.")
        if request.source_text != row.original_text:
            raise ValueError(f"Translation result source text does not match row: {row.request_id}.")
        matched.append((row, result))
    return tuple(matched)


def _write_review_header(worksheet, source_name: str, source_language: str, target_language: str) -> None:
    """Reuse the established translation review structure and styling."""

    worksheet.merge_cells("A1:K1")
    title = worksheet.cell(row=1, column=1, value=TITLE_TEXT)
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24
    metadata = (("Video:", source_name), ("Date (Processed):", ""), ("Date (Translated):", ""),
                ("Translator(s):", ""), ("Language (Source):", source_language),
                ("Language (Target):", target_language), ("Notes:", ""))
    for offset, (label, value) in enumerate(metadata, start=2):
        worksheet.cell(row=offset, column=1, value=label).font = Font(bold=True)
        worksheet.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=11)
        cell = worksheet.cell(row=offset, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.protection = Protection(locked=False)
    worksheet.row_dimensions[8].height = 36
    for column, header in enumerate(_TABLE_HEADERS, start=1):
        cell = worksheet.cell(row=TABLE_HEADER_ROW, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[TABLE_HEADER_ROW].height = HEADER_ROW_HEIGHT
    worksheet.cell(row=9, column=1, value=_REVIEW_GUIDANCE).alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells("A9:K9")
    worksheet.row_dimensions[9].height = 30
    worksheet.freeze_panes = f"A{TABLE_START_ROW}"


def _write_sheet(worksheet, sheet: PlannedSheet, source_name: str, source_language: str,
                 matches: tuple[tuple[TranslationWorkbookRow, TranslationResult], ...],
                 assessments: Mapping[str, TranslationReviewAssessment],
                 human_reviews: Mapping[str, HumanTranslationReview]) -> list[dict[str, Any]]:
    """Write one review sheet and return non-user-facing row provenance."""

    _write_review_header(worksheet, source_name, source_language, sheet.target_language)
    metadata_rows: list[dict[str, Any]] = []
    for row, result in matches:
        number = worksheet.max_row + 1
        translation = result.translated_text if result.status is TranslationStatus.SUCCESS else None
        assessment = assessments[result.request_id]
        human_review = human_reviews[result.request_id]
        resolution = resolve_reviewed_translation(result.status, result.translated_text, human_review)
        verified_translation = resolution.output_text if resolution.human_verified else None
        review_reasons = "; ".join(warning.explanation for warning in assessment.warnings)
        cells = (
            worksheet.cell(number, 1, row.slide_value), worksheet.cell(number, 2, row.original_text),
            worksheet.cell(number, 3, translation), worksheet.cell(number, 4, verified_translation),
            worksheet.cell(number, 5, human_review_status_display(human_review.status)),
            worksheet.cell(number, 6, review_status_display(assessment.status)),
            worksheet.cell(number, 7, review_reasons or None),
            worksheet.cell(number, 8, human_review.reviewer_notes),
            worksheet.cell(number, 9, result.request.target_language),
            worksheet.cell(number, 10, result.provider_id), worksheet.cell(number, 11, result.model_id),
        )
        cells[0].protection = Protection(locked=True); cells[0].fill = SOURCE_REFERENCE_FILL
        cells[1].alignment = Alignment(wrap_text=True, vertical="top"); cells[1].protection = Protection(locked=True); cells[1].fill = SOURCE_REFERENCE_FILL
        cells[2].alignment = Alignment(wrap_text=True, vertical="top"); cells[2].protection = Protection(locked=True); cells[2].fill = AI_TRANSLATION_FILL
        cells[3].alignment = Alignment(wrap_text=True, vertical="top"); cells[3].protection = Protection(locked=False); cells[3].fill = MODIFIED_TRANSLATION_FILL
        cells[3].comment = Comment(_VERIFIED_TRANSLATION_NOTE, "VideoText")
        cells[4].protection = Protection(locked=False); cells[4].fill = VERIFIED_FILL
        cells[5].alignment = Alignment(wrap_text=True, vertical="top"); cells[5].protection = Protection(locked=True)
        cells[6].alignment = Alignment(wrap_text=True, vertical="top"); cells[6].protection = Protection(locked=True)
        cells[7].alignment = Alignment(wrap_text=True, vertical="top"); cells[7].protection = Protection(locked=False)
        for cell in cells[8:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.protection = Protection(locked=True)
        if assessment.warnings:
            cells[5].comment = Comment(
                "Review warnings:\n" + "\n".join(f"- {warning.explanation}" for warning in assessment.warnings),
                "VideoText",
            )
        worksheet.row_dimensions[number].height = _estimate_row_height(
            row.original_text, translation or "", verified_translation or "",
            review_reasons, human_review.reviewer_notes or "",
        )
        metadata_rows.append({
            "sheet_id": sheet.sheet_id, "source_item_id": row.source_item_id,
            "source_reference": row.source_reference, "request_id": result.request_id,
            "source_type": result.request.provenance.source_type.value,
            "target_language": result.request.target_language, "provider_id": result.provider_id,
            "model_id": result.model_id, "status": result.status.value,
            "review_status": assessment.status.value,
            "review_revision": assessment.revision,
            "review_warnings": json.dumps([warning.code.value for warning in assessment.warnings]),
            "review_warning_context": json.dumps(
                {warning.code.value: dict(warning.context) for warning in assessment.warnings},
                ensure_ascii=False, sort_keys=True),
            "error": _safe_error(result.error), "ordering_index": row.ordering_index,
            "provider_metadata": json.dumps(_safe_metadata(result.provider_metadata), ensure_ascii=False, sort_keys=True),
        })
    worksheet.column_dimensions["A"].width = 10
    for column in ("B", "C", "D", "G", "H"):
        worksheet.column_dimensions[column].width = TEXT_COLUMN_WIDTH
    worksheet.column_dimensions["E"].width = VERIFIED_COLUMN_WIDTH
    worksheet.column_dimensions["F"].width = 24
    for column in ("I", "J", "K"):
        worksheet.column_dimensions[column].width = 20
    worksheet.auto_filter.ref = f"A{TABLE_HEADER_ROW}:K{worksheet.max_row}"
    validation = DataValidation(
        type="list", formula1='"Unreviewed,Accepted,Edited / Verified,Flagged"', allow_blank=False)
    validation.promptTitle = "Human Review Status"
    validation.prompt = "Select the human review disposition for this translation."
    validation.showInputMessage = True
    worksheet.add_data_validation(validation)
    if worksheet.max_row >= TABLE_START_ROW:
        validation.add(f"E{TABLE_START_ROW}:E{worksheet.max_row}")
    worksheet.protection.sheet = True
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False
    worksheet.protection.formatColumns = False
    worksheet.protection.autoFilter = False
    return metadata_rows


def _write_metadata_sheet(workbook: Workbook, job: TranslationJob, workbook_id: str, rows: list[dict[str, Any]]) -> None:
    """Store deterministic, non-secret provenance outside the review table."""

    sheet = workbook.create_sheet(_METADATA_SHEET_NAME)
    sheet.append(("VideoText Translation Schema", _SCHEMA_VERSION))
    for key, value in (("job_id", job.job_id), ("workbook_id", workbook_id),
                       ("grouping_mode", job.output_plan.grouping.value), ("provider_name", job.provider_name)):
        sheet.append((key, value))
    headers = ("sheet_id", "source_item_id", "source_reference", "request_id", "source_type",
               "target_language", "provider_id", "model_id", "status", "review_status", "review_revision",
               "review_warnings", "review_warning_context", "error", "ordering_index", "provider_metadata")
    sheet.append(())
    sheet.append(headers)
    for row in rows:
        sheet.append(tuple(row[header] for header in headers))
    sheet.sheet_state = "hidden"


def _save_new_workbook(workbook: Workbook, final_path: Path) -> None:
    """Save through a temporary sibling and never overwrite an existing target."""

    if final_path.exists():
        raise FileExistsError(f"Translation workbook already exists: {final_path}")
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{final_path.stem}-", suffix=".xlsx", dir=final_path.parent)
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        try:
            with final_path.open("xb"):
                pass
        except FileExistsError as error:
            raise FileExistsError(f"Translation workbook already exists: {final_path}") from error
        os.replace(temporary_path, final_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def populate_translation_workbooks(job: TranslationJob, output_layout: TranslationOutputLayout,
                                   source_rows: Iterable[TranslationWorkbookRow], translation_results: Iterable[TranslationResult],
                                   output_directory: str | Path,
                                   assessments: Iterable[TranslationReviewAssessment] | None = None,
                                   human_reviews: Iterable[HumanTranslationReview] | None = None) -> TranslationWorkbookWriteResult:
    """Create only new planned workbooks from complete, already-decided evidence."""

    if not isinstance(job, TranslationJob) or not isinstance(output_layout, TranslationOutputLayout):
        raise ValueError("job and output_layout must use translation planning models.")
    if output_layout.job_id != job.job_id:
        raise ValueError("output_layout does not belong to the translation job.")
    if not isinstance(output_directory, (str, Path)) or not str(output_directory).strip():
        raise ValueError("output_directory is required.")
    directory = Path(output_directory).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    rows_by_source = _rows_by_source(source_rows, job)
    results = _result_map(translation_results)
    assessment_values = tuple(assessments) if assessments is not None else assess_translation_results(results.values())
    assessment_map = {assessment.request_id: assessment for assessment in assessment_values}
    if len(assessment_map) != len(assessment_values) or set(assessment_map) != set(results):
        raise ValueError("Translation review assessments must match translation results exactly.")
    human_review_map = _human_review_map(human_reviews, set(results))
    source_names = {item.source_item_id: item.display_name for item in job.source_items}
    prepared: list[tuple[Any, tuple[tuple[PlannedSheet, tuple[tuple[TranslationWorkbookRow, TranslationResult], ...]], ...]]] = []
    expected_ids: set[str] = set()
    for planned_workbook in output_layout.workbooks:
        sheets = []
        for sheet in planned_workbook.sheets:
            matches = _validate_sheet_evidence(sheet, rows_by_source[sheet.source_item_id], results)
            expected_ids.update(row.request_id for row, _ in matches)
            sheets.append((sheet, matches))
        prepared.append((planned_workbook, tuple(sheets)))
    extra = set(results) - expected_ids
    if extra:
        raise ValueError(f"Translation result does not belong to the planned output: {sorted(extra)[0]}.")
    paths: dict[str, Path] = {}
    success_count = failure_count = row_count = sheet_count = 0
    for planned_workbook, sheets in prepared:
        final_path = directory / planned_workbook.filename
        if final_path.parent != directory:
            raise ValueError("Planned workbook filename escapes output_directory.")
        workbook = Workbook()
        metadata_rows: list[dict[str, Any]] = []
        first = True
        for sheet, matches in sheets:
            worksheet = workbook.active if first else workbook.create_sheet()
            first = False
            worksheet.title = sheet.name
            metadata_rows.extend(_write_sheet(worksheet, sheet, source_names[sheet.source_item_id], job.source_language,
                                              matches, assessment_map, human_review_map))
            row_count += len(matches); sheet_count += 1
            success_count += sum(result.status is TranslationStatus.SUCCESS for _, result in matches)
            failure_count += sum(result.status is TranslationStatus.FAILURE for _, result in matches)
        _write_metadata_sheet(workbook, job, planned_workbook.workbook_id, metadata_rows)
        _save_new_workbook(workbook, final_path)
        paths[planned_workbook.workbook_id] = final_path
    return TranslationWorkbookWriteResult(paths, sheet_count, row_count, success_count, failure_count)
