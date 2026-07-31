"""
excel_exporter.py

Exports reconstructed presentations to Excel.
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from math import ceil
from pathlib import Path
import re

from models import Presentation


BASE_ROW_HEIGHT = 15
MIN_CONTENT_ROW_HEIGHT = 15
MAX_CONTENT_ROW_HEIGHT = 300
TEXT_COLUMN_WIDTH = 60
VERIFIED_COLUMN_WIDTH = 12
HEADER_ROW_HEIGHT = 20
TITLE_ROW = 1
METADATA_START_ROW = 2
TABLE_HEADER_ROW = 10
TABLE_START_ROW = TABLE_HEADER_ROW + 1
TITLE_TEXT = "VideoText Translation Workbook"
TABLE_HEADERS = (
    "Slide",
    "Original Text",
    "Initial AI Translation",
    "Modified Translation",
    "Verified",
)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SOURCE_REFERENCE_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
AI_TRANSLATION_FILL = PatternFill(fill_type="solid", fgColor="EAF1FB")
MODIFIED_TRANSLATION_FILL = PatternFill(fill_type="solid", fgColor="EAF6EA")
VERIFIED_FILL = PatternFill(fill_type="solid", fgColor="FFF6D8")
MODIFIED_TRANSLATION_NOTE = "Make edits to the AI Translation here if needed."
VERIFIED_INPUT_MESSAGE = "Verify translation is accurate (Yes / No)."


def _format_paragraph_text(paragraph) -> tuple[str, int]:
    """Return canonical paragraph prose without inventing list markers.

    Paragraph reconstruction has already preserved any genuine bullet or
    numbered marker in ``paragraph.text``.  Excel should wrap ordinary prose
    naturally rather than turning visual continuation lines into new bullets.
    """

    return paragraph.text, max(1, len(paragraph.text.splitlines()))


def _estimate_wrapped_lines(text: str, column_width: float) -> int:
    """Estimate visible Excel lines from explicit breaks and bounded width."""
    characters_per_line = max(1, int(column_width))
    lines = text.splitlines() or [""]
    return sum(max(1, ceil(len(line) / characters_per_line)) for line in lines)


def _estimate_row_height(*cell_texts: str) -> float:
    """Return a bounded height for the most demanding wrapped cell in a row."""
    visible_lines = max(
        _estimate_wrapped_lines(text, TEXT_COLUMN_WIDTH)
        for text in cell_texts
    )
    return min(
        MAX_CONTENT_ROW_HEIGHT,
        max(MIN_CONTENT_ROW_HEIGHT, BASE_ROW_HEIGHT * visible_lines),
    )


def _video_name(presentation: Presentation) -> str:
    """Return the processed video stem when metadata makes it available."""

    video_path = presentation.metadata.get("video_path")
    if video_path:
        return Path(str(video_path)).stem

    checkpoint_path = presentation.metadata.get("source_checkpoint")
    if checkpoint_path:
        checkpoint = Path(str(checkpoint_path))
        if checkpoint.parent.name.lower() == "cache":
            return checkpoint.parent.parent.name

    return ""


def _write_metadata(worksheet, presentation: Presentation) -> None:
    """Write editable translation metadata above the slide table."""

    worksheet.merge_cells("A1:E1")
    title_cell = worksheet.cell(row=TITLE_ROW, column=1, value=TITLE_TEXT)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[TITLE_ROW].height = 24

    metadata = (
        ("Video:", _video_name(presentation)),
        ("Date (Processed):", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Date (Translated):", ""),
        ("Translator(s):", ""),
        ("Language (Source):", ""),
        ("Language (Target):", ""),
        ("Notes:", ""),
    )

    for offset, (label, value) in enumerate(metadata):
        row = METADATA_START_ROW + offset
        worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        value_cell = worksheet.cell(row=row, column=2, value=value)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Preserve the existing editable metadata fields outside the table.
        value_cell.protection = Protection(locked=False)

    worksheet.row_dimensions[METADATA_START_ROW + len(metadata) - 1].height = 36


def sanitize_worksheet_name(name: str, used_names: set[str] | None = None) -> str:
    """Return a deterministic, unique Excel worksheet name within 31 chars."""

    used_names = used_names if used_names is not None else set()
    base = re.sub(r"[:\\\\/?*\[\]]", "_", name).strip()
    base = base or "Video"
    base = base[:31]
    candidate = base
    suffix_number = 2
    while candidate.casefold() in used_names:
        suffix = f" ({suffix_number})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        suffix_number += 1
    used_names.add(candidate.casefold())
    return candidate


def write_presentation_worksheet(worksheet, presentation: Presentation) -> None:
    """Write one translation-ready presentation worksheet with existing styles."""

    _write_metadata(worksheet, presentation)

    for column, header in enumerate(TABLE_HEADERS, start=1):
        cell = worksheet.cell(row=TABLE_HEADER_ROW, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    worksheet.row_dimensions[TABLE_HEADER_ROW].height = HEADER_ROW_HEIGHT
    worksheet.freeze_panes = f"A{TABLE_START_ROW}"

    for slide in presentation.slides:
        for paragraph in slide.paragraphs:
            paragraph_text, _line_count = _format_paragraph_text(paragraph)
            row_number = worksheet.max_row + 1
            slide_cell = worksheet.cell(row=row_number, column=1, value=slide.slide_number)
            slide_cell.protection = Protection(locked=True)
            slide_cell.fill = SOURCE_REFERENCE_FILL
            original_text_cell = worksheet.cell(row=row_number, column=2, value=paragraph_text)
            original_text_cell.alignment = Alignment(wrap_text=True, vertical="top")
            original_text_cell.protection = Protection(locked=True)
            original_text_cell.fill = SOURCE_REFERENCE_FILL

            ai_translation_cell = worksheet.cell(row=row_number, column=3, value="")
            ai_translation_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ai_translation_cell.protection = Protection(locked=True)
            ai_translation_cell.fill = AI_TRANSLATION_FILL

            modified_translation_cell = worksheet.cell(row=row_number, column=4, value="")
            modified_translation_cell.alignment = Alignment(wrap_text=True, vertical="top")
            modified_translation_cell.protection = Protection(locked=False)
            modified_translation_cell.fill = MODIFIED_TRANSLATION_FILL
            modified_translation_cell.comment = Comment(
                MODIFIED_TRANSLATION_NOTE,
                "VideoText",
            )

            verified_cell = worksheet.cell(row=row_number, column=5, value="")
            verified_cell.protection = Protection(locked=False)
            verified_cell.fill = VERIFIED_FILL
            worksheet.row_dimensions[row_number].height = _estimate_row_height(
                paragraph_text,
                "",
                "",
            )

    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = TEXT_COLUMN_WIDTH
    worksheet.column_dimensions["C"].width = TEXT_COLUMN_WIDTH
    worksheet.column_dimensions["D"].width = TEXT_COLUMN_WIDTH
    worksheet.column_dimensions["E"].width = VERIFIED_COLUMN_WIDTH
    worksheet.auto_filter.ref = f"A{TABLE_HEADER_ROW}:E{worksheet.max_row}"

    verification = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
    )
    verification.promptTitle = "Verified"
    verification.prompt = VERIFIED_INPUT_MESSAGE
    verification.showInputMessage = True
    worksheet.add_data_validation(verification)
    if worksheet.max_row >= TABLE_START_ROW:
        verification.add(f"E{TABLE_START_ROW}:E{worksheet.max_row}")

    # No password is set: protected cells stay read-only while translators can
    # edit columns D/E, select editable cells, resize columns, and filter.
    worksheet.protection.sheet = True
    # Sheet-protection flags use ``True`` to prohibit an action.  Keep both
    # selections permitted so Excel can select—and therefore edit—the
    # explicitly unlocked translation and verification cells.
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False
    worksheet.protection.formatColumns = False
    worksheet.protection.autoFilter = False


def export_excel(
    presentation: Presentation,
    output_path: str,
) -> str:
    """
    Export one worksheet row for each paragraph in presentation order.
    """

    workbook = Workbook()
    worksheet = workbook.active

    write_presentation_worksheet(worksheet, presentation)

    workbook.save(output_path)

    return output_path
