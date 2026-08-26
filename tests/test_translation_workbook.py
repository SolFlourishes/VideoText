"""Round-trip tests for planned, evidence-only translation review workbooks."""

from __future__ import annotations

from pathlib import Path
import sys
import tempfile
import unittest

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from translation_contract import (TranslationProvenance, TranslationRequest, TranslationResult,
    TranslationSourceType, TranslationStatus)
from translation_job import TranslationJob, TranslationOutputGrouping, TranslationOutputPlan, TranslationSourceItem
from translation_output_plan import plan_translation_output
from translation_review import (HumanTranslationReview, HumanTranslationReviewStatus,
    TranslationReviewAssessment, TranslationReviewStatus, TranslationReviewWarning,
    TranslationReviewWarningCode)
from translation_workbook import TranslationWorkbookRow, populate_translation_workbooks


def make_job(grouping: TranslationOutputGrouping, languages: tuple[str, ...] = ("es",), sources: tuple[str, ...] = ("Video A",)) -> TranslationJob:
    items = tuple(TranslationSourceItem(f"video-{number}", name, f"project:{number}", number) for number, name in enumerate(sources))
    return TranslationJob("job-38f2", items, "en", languages, "argos", TranslationOutputPlan(grouping))


def evidence(job: TranslationJob, failed: bool = False):
    rows, results = [], []
    for item in job.source_items:
        for index, text in enumerate(("First source line", "Second\nsource line")):
            reference = f"{item.source_item_id}:slide:1:paragraph:{index}"
            for language in job.target_languages:
                request = TranslationRequest(f"{reference}:translation:{language}", text, "en", language,
                    TranslationProvenance(reference, TranslationSourceType.VERIFIED_OCR), index)
                rows.append(TranslationWorkbookRow(item.source_item_id, reference, request.request_id, language, 1, text, index))
                if failed and index == 1:
                    results.append(TranslationResult(request, TranslationStatus.FAILURE, "argos", error="Network token=secret unavailable"))
                else:
                    results.append(TranslationResult(request, TranslationStatus.SUCCESS, "argos", f"{language}: {text}", model_id="model-1", provider_metadata={"quality": "test", "api_key": "not-written"}))
    return tuple(rows), tuple(results)


class TranslationWorkbookTests(unittest.TestCase):
    def write(self, job: TranslationJob, rows=None, results=None):
        rows, results = evidence(job) if rows is None else (rows, results)
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        result = populate_translation_workbooks(job, plan_translation_output(job), rows, results, temporary.name)
        return result, Path(temporary.name)

    def test_basic_workbook_exposes_filterable_review_evidence_without_overwriting_sources(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        result, _directory = self.write(job)
        path = next(iter(result.workbook_paths.values()))
        book = load_workbook(path)
        sheet = book["Video A"]
        self.assertEqual([sheet.cell(10, column).value for column in range(1, 12)], [
            "Slide", "Source OCR Text", "Original AI Translation", "Verified Translation",
            "Human Review Status", "Translation Review Status", "Review Reasons",
            "Reviewer Notes", "Target Language", "Provider", "Model"])
        self.assertEqual("Normal Review", sheet["F11"].value)
        self.assertIn("AI-generated translations require human review", sheet["A9"].value)
        self.assertEqual(1, sheet["A11"].value)
        self.assertEqual("First source line", sheet["B11"].value)
        self.assertEqual("es: First source line", sheet["C11"].value)
        self.assertIsNone(sheet["D11"].value); self.assertEqual("Unreviewed", sheet["E11"].value)
        self.assertIsNone(sheet["G11"].value); self.assertIsNone(sheet["H11"].value)
        self.assertEqual("es", sheet["I11"].value); self.assertEqual("argos", sheet["J11"].value)
        self.assertEqual("model-1", sheet["K11"].value)
        self.assertTrue(sheet["B11"].protection.locked); self.assertTrue(sheet["C11"].protection.locked)
        self.assertFalse(sheet["D11"].protection.locked); self.assertFalse(sheet["E11"].protection.locked)
        self.assertFalse(sheet["H11"].protection.locked); self.assertTrue(sheet["G11"].protection.locked)
        self.assertTrue(sheet["B12"].alignment.wrap_text); self.assertIn("E11", str(list(sheet.data_validations.dataValidation)[0].sqref))
        self.assertEqual('\"Unreviewed,Accepted,Edited / Verified,Flagged\"',
                         list(sheet.data_validations.dataValidation)[0].formula1)
        self.assertEqual("A10:K12", sheet.auto_filter.ref)
        metadata = book["_VideoText_Metadata"]
        self.assertEqual("hidden", metadata.sheet_state)
        self.assertEqual("VideoText Translation Schema", metadata["A1"].value)
        self.assertEqual("2", metadata["B1"].value)

    def test_existing_human_review_populates_verified_text_status_and_notes(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job)
        reviews = (
            HumanTranslationReview(results[0].request_id, HumanTranslationReviewStatus.ACCEPTED,
                                   reviewer_notes="AI translation accepted"),
            HumanTranslationReview(results[1].request_id, HumanTranslationReviewStatus.EDITED_VERIFIED,
                                   "Human revised translation", "Terminology corrected"),
        )
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        written = populate_translation_workbooks(
            job, plan_translation_output(job), rows, results, temporary.name, human_reviews=reviews)
        sheet = load_workbook(next(iter(written.workbook_paths.values())))["Video A"]
        self.assertEqual(sheet["C11"].value, sheet["D11"].value)
        self.assertEqual("Accepted", sheet["E11"].value)
        self.assertEqual("AI translation accepted", sheet["H11"].value)
        self.assertEqual("Human revised translation", sheet["D12"].value)
        self.assertEqual("Edited / Verified", sheet["E12"].value)
        self.assertEqual("Terminology corrected", sheet["H12"].value)
        self.assertEqual("es: Second\nsource line", sheet["C12"].value)

    def test_flagged_review_preserves_ai_text_and_is_filterable(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job)
        reviews = (
            HumanTranslationReview(
                results[0].request_id, HumanTranslationReviewStatus.FLAGGED,
                reviewer_notes="Needs specialist review"),
            HumanTranslationReview(results[1].request_id),
        )
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        written = populate_translation_workbooks(
            job, plan_translation_output(job), rows, results, temporary.name, human_reviews=reviews)
        sheet = load_workbook(next(iter(written.workbook_paths.values())))["Video A"]
        self.assertEqual("es: First source line", sheet["C11"].value)
        self.assertIsNone(sheet["D11"].value)
        self.assertEqual("Flagged", sheet["E11"].value)
        self.assertEqual("Needs specialist review", sheet["H11"].value)
        self.assertEqual("A10:K12", sheet.auto_filter.ref)

    def test_all_grouping_modes_use_planned_names_and_order(self) -> None:
        cases = ((TranslationOutputGrouping.BY_LANGUAGE, ("es", "de"), ("Video A", "Video B"), 2, ("Spanish.xlsx", "German.xlsx")),
                 (TranslationOutputGrouping.BY_SOURCE, ("es", "de"), ("Video A",), 1, ("Video A.xlsx",)),
                 (TranslationOutputGrouping.COMBINED, ("es", "de"), ("Video A", "Video B"), 1, ("Translation Batch.xlsx",)),
                 (TranslationOutputGrouping.SEPARATE, ("es", "de"), ("Video A",), 2, ("Video A - Spanish.xlsx", "Video A - German.xlsx")))
        for grouping, languages, sources, workbook_count, filenames in cases:
            with self.subTest(grouping=grouping):
                job = make_job(grouping, languages, sources)
                result, _directory = self.write(job)
                self.assertEqual(workbook_count, len(result.workbook_paths))
                self.assertEqual(filenames, tuple(path.name for path in result.workbook_paths.values()))
                for workbook_id, path in result.workbook_paths.items():
                    expected = next(book for book in plan_translation_output(job).workbooks if book.workbook_id == workbook_id)
                    book = load_workbook(path)
                    self.assertEqual(tuple(sheet.name for sheet in expected.sheets), tuple(book.sheetnames[:-1]))
                    for sheet in book.worksheets[:-1]:
                        self.assertEqual("Human Review Status", sheet["E10"].value)
                        self.assertEqual("Translation Review Status", sheet["F10"].value)
                        self.assertEqual(f"A10:K{sheet.max_row}", sheet.auto_filter.ref)

    def test_failure_is_blank_in_review_table_and_auditable_in_hidden_metadata(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job, failed=True)
        result, _directory = self.write(job, rows, results)
        book = load_workbook(next(iter(result.workbook_paths.values())))
        sheet, metadata = book["Video A"], book["_VideoText_Metadata"]
        self.assertIsNone(sheet["C12"].value)
        self.assertEqual("Translation Failed", sheet["F12"].value)
        self.assertEqual("No usable provider translation was produced.", sheet["G12"].value)
        headers = [metadata.cell(7, column).value for column in range(1, 17)]
        values = [metadata.cell(9, column).value for column in range(1, 17)]
        record = dict(zip(headers, values))
        self.assertEqual("failure", record["status"])
        self.assertEqual("argos", record["provider_id"])
        self.assertEqual("verified_ocr", record["source_type"])
        self.assertNotIn("secret", record["error"])
        self.assertEqual("translation_failed", record["review_status"])
        self.assertEqual('["translation_failed"]', record["review_warnings"])
        self.assertIn("No usable provider translation", sheet["F12"].comment.text)
        self.assertEqual(1, result.failure_translation_count)

    def test_review_metadata_cannot_promote_failed_translation_in_workbook(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job, failed=True)
        reviews = (
            HumanTranslationReview(results[0].request_id),
            HumanTranslationReview(
                results[1].request_id, HumanTranslationReviewStatus.EDITED_VERIFIED,
                "Human text attached to failed provider evidence"),
        )
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        written = populate_translation_workbooks(
            job, plan_translation_output(job), rows, results, temporary.name, human_reviews=reviews)
        sheet = load_workbook(next(iter(written.workbook_paths.values())))["Video A"]
        self.assertIsNone(sheet["C12"].value)
        self.assertIsNone(sheet["D12"].value)
        self.assertEqual("Edited / Verified", sheet["E12"].value)
        self.assertEqual("Translation Failed", sheet["F12"].value)

    def test_review_recommended_status_and_reason_are_visible_and_filterable(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job)
        warning = TranslationReviewWarning(
            TranslationReviewWarningCode.NUMERIC_MISMATCH,
            "Numeric tokens differ between source and translation.",
            {"source_numbers": ("25",), "translation_numbers": ("24",)},
        )
        assessments = (
            TranslationReviewAssessment(
                results[0].request_id, TranslationReviewStatus.REVIEW_RECOMMENDED, (warning,)),
            TranslationReviewAssessment(
                results[1].request_id, TranslationReviewStatus.NORMAL_REVIEW),
        )
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        written = populate_translation_workbooks(
            job, plan_translation_output(job), rows, results, temporary.name, assessments=assessments)
        sheet = load_workbook(next(iter(written.workbook_paths.values())))["Video A"]
        self.assertEqual("Review Recommended", sheet["F11"].value)
        self.assertEqual("Numeric tokens differ between source and translation.", sheet["G11"].value)
        self.assertEqual("A10:K12", sheet.auto_filter.ref)

    def test_missing_duplicate_and_mismatched_results_fail_strictly(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job)
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(ValueError, "Missing translation result"):
                populate_translation_workbooks(job, plan_translation_output(job), rows, results[:-1], directory)
            with self.assertRaisesRegex(ValueError, "Duplicate translation result"):
                populate_translation_workbooks(job, plan_translation_output(job), rows, results + (results[0],), directory)
            mismatched = TranslationResult(
                TranslationRequest(results[0].request_id, results[0].source_text, "en", "de",
                    results[0].request.provenance, 0), TranslationStatus.SUCCESS, "argos", "Deutsch")
            with self.assertRaisesRegex(ValueError, "target language"):
                populate_translation_workbooks(job, plan_translation_output(job), rows, (mismatched,) + results[1:], directory)
            wrong_reference = TranslationResult(
                TranslationRequest(results[0].request_id, results[0].source_text, "en", "es",
                    TranslationProvenance("different:source", TranslationSourceType.OCR), 0),
                TranslationStatus.SUCCESS, "argos", "Traduccion")
            with self.assertRaisesRegex(ValueError, "source reference"):
                populate_translation_workbooks(job, plan_translation_output(job), rows, (wrong_reference,) + results[1:], directory)

    def test_existing_target_is_never_overwritten_and_paths_stay_in_directory(self) -> None:
        job = make_job(TranslationOutputGrouping.BY_LANGUAGE)
        rows, results = evidence(job)
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "Spanish.xlsx"
            target.write_bytes(b"existing review work")
            with self.assertRaisesRegex(FileExistsError, "already exists"):
                populate_translation_workbooks(job, plan_translation_output(job), rows, results, directory)
            self.assertEqual(b"existing review work", target.read_bytes())
