"""Focused tests for the optional consolidated batch translation workbook."""

from pathlib import Path
import sys
import tempfile
import unittest

from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from batch_excel_exporter import export_batch_excel
from excel_exporter import sanitize_worksheet_name
from models import Presentation, Slide, TextParagraph, TextType


def presentation(name: str) -> Presentation:
    return Presentation(
        metadata={"video_path": name},
        slides=[Slide(
            slide_number=1,
            start_time=0,
            end_time=1,
            paragraphs=[TextParagraph("Extracted text", text_type=TextType.BODY)],
        )],
    )


class BatchExcelExporterTests(unittest.TestCase):
    def test_worksheet_names_are_safe_unique_and_bounded(self):
        used: set[str] = set()
        self.assertEqual(sanitize_worksheet_name("Week 1: Overview", used), "Week 1_ Overview")
        self.assertEqual(sanitize_worksheet_name("Lecture", used), "Lecture")
        self.assertEqual(sanitize_worksheet_name("lecture", used), "lecture (2)")
        self.assertEqual(
            sanitize_worksheet_name("Introduction to Clinical Education and Translation", used),
            "Introduction to Clinical Educat",
        )

    def test_consolidated_workbook_has_one_formatted_sheet_per_video(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            destination = Path(temporary_directory) / "Batch VideoText Export.xlsx"
            result = export_batch_excel([
                ("Week 1: Overview.mp4", presentation("Week 1: Overview.mp4")),
                ("Week 1? Overview.avi", presentation("Week 1? Overview.avi")),
            ], destination)
            workbook = load_workbook(result)

        self.assertEqual(workbook.sheetnames, ["Week 1_ Overview", "Week 1_ Overview (2)"])
        for worksheet in workbook.worksheets:
            self.assertEqual(worksheet["A1"].value, "VideoText Translation Workbook")
            self.assertEqual(worksheet["B11"].value, "Extracted text")
            self.assertEqual(worksheet.freeze_panes, "A11")
            self.assertEqual(worksheet.auto_filter.ref, "A10:E11")

    def test_empty_presentations_do_not_create_a_workbook(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            destination = Path(temporary_directory) / "Batch VideoText Export.xlsx"
            with self.assertRaisesRegex(ValueError, "at least one successful"):
                export_batch_excel([], destination)
            self.assertFalse(destination.exists())


if __name__ == "__main__":
    unittest.main()
