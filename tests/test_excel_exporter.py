"""Focused tests for the VideoText translation workbook layout."""

import csv
import re
import sys
import tempfile
from pathlib import Path
import unittest

from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from csv_exporter import export_csv
from excel_exporter import (
    AI_TRANSLATION_FILL,
    MAX_CONTENT_ROW_HEIGHT,
    MODIFIED_TRANSLATION_NOTE,
    MODIFIED_TRANSLATION_FILL,
    SOURCE_REFERENCE_FILL,
    VERIFIED_INPUT_MESSAGE,
    VERIFIED_FILL,
    export_excel,
)
from markdown_exporter import export_markdown
from models import Presentation, Slide, TextLine, TextParagraph, TextType


def text_line(text: str) -> TextLine:
    return TextLine(
        text=text,
        top=0,
        bottom=1,
        left=0,
        right=1,
        confidence=1.0,
    )


class ExcelExporterTests(unittest.TestCase):
    def export_workbook(self, paragraphs: list[TextParagraph]):
        presentation = Presentation(
            metadata={"video_path": "D:/Videos/sample.mp4"},
            slides=[
                Slide(
                    slide_number=7,
                    start_time=0.0,
                    end_time=1.0,
                    paragraphs=paragraphs,
                ),
            ],
        )
        temporary_directory = tempfile.TemporaryDirectory()
        output_path = Path(temporary_directory.name) / "export.xlsx"
        export_excel(presentation, str(output_path))
        self.addCleanup(temporary_directory.cleanup)
        return load_workbook(output_path).active

    def test_title_and_editable_metadata_fields(self):
        worksheet = self.export_workbook([])

        self.assertEqual(worksheet["A1"].value, "VideoText Translation Workbook")
        self.assertEqual(
            {str(cell_range) for cell_range in worksheet.merged_cells.ranges},
            {"A1:E1", "B2:E2", "B3:E3", "B4:E4", "B5:E5", "B6:E6", "B7:E7", "B8:E8"},
        )
        self.assertEqual(worksheet["A2"].value, "Video:")
        self.assertEqual(worksheet["B2"].value, "sample")
        self.assertRegex(worksheet["B3"].value, r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$")
        self.assertIsNone(worksheet["B4"].value)
        self.assertIsNone(worksheet["B5"].value)
        self.assertIsNone(worksheet["B6"].value)
        self.assertIsNone(worksheet["B7"].value)
        self.assertIsNone(worksheet["B8"].value)

    def test_table_headers_freeze_panes_and_filter(self):
        worksheet = self.export_workbook([])

        self.assertEqual(
            [worksheet.cell(row=10, column=column).value for column in range(1, 6)],
            [
                "Slide",
                "Original Text",
                "Initial AI Translation",
                "Modified Translation",
                "Verified",
            ],
        )
        self.assertEqual(worksheet.freeze_panes, "A11")
        self.assertEqual(worksheet.auto_filter.ref, "A10:E10")
        self.assertTrue(worksheet["A10"].font.bold)
        self.assertTrue(worksheet["A10"].alignment.wrap_text)
        self.assertEqual(worksheet["A10"].alignment.vertical, "center")
        self.assertLessEqual(worksheet.row_dimensions[10].height, 20)

    def test_original_text_is_preserved_and_translation_workflow_is_blank(self):
        original = "One line"
        worksheet = self.export_workbook([
            TextParagraph(original, text_type=TextType.BODY),
        ])

        self.assertEqual(worksheet["A11"].value, 7)
        self.assertEqual(worksheet["B11"].value, original)
        self.assertIsNone(worksheet["C11"].value)
        self.assertIsNone(worksheet["D11"].value)
        self.assertIsNone(worksheet["E11"].value)
        self.assertTrue(worksheet["A11"].protection.locked)
        self.assertTrue(worksheet["B11"].protection.locked)
        self.assertTrue(worksheet["C11"].protection.locked)
        self.assertFalse(worksheet["D11"].protection.locked)
        self.assertFalse(worksheet["E11"].protection.locked)
        self.assertTrue(worksheet.protection.sheet)
        self.assertFalse(worksheet.protection.selectUnlockedCells)
        self.assertEqual(worksheet["A11"].fill.fgColor.rgb, SOURCE_REFERENCE_FILL.fgColor.rgb)
        self.assertEqual(worksheet["B11"].fill.fgColor.rgb, SOURCE_REFERENCE_FILL.fgColor.rgb)
        self.assertEqual(worksheet["C11"].fill.fgColor.rgb, AI_TRANSLATION_FILL.fgColor.rgb)
        self.assertEqual(worksheet["D11"].fill.fgColor.rgb, MODIFIED_TRANSLATION_FILL.fgColor.rgb)
        self.assertEqual(worksheet["E11"].fill.fgColor.rgb, VERIFIED_FILL.fgColor.rgb)
        self.assertTrue(worksheet["B11"].alignment.wrap_text)
        self.assertEqual(worksheet["B11"].alignment.vertical, "top")
        self.assertTrue(worksheet["C11"].alignment.wrap_text)
        self.assertEqual(worksheet["C11"].alignment.vertical, "top")
        self.assertTrue(worksheet["D11"].alignment.wrap_text)
        self.assertEqual(worksheet["D11"].alignment.vertical, "top")

    def test_translation_guidance_and_verification_validation_are_present(self):
        worksheet = self.export_workbook([
            TextParagraph("Original", text_type=TextType.BODY),
        ])

        self.assertEqual(worksheet["D11"].comment.text, MODIFIED_TRANSLATION_NOTE)
        validations = list(worksheet.data_validations.dataValidation)
        self.assertEqual(len(validations), 1)
        validation = validations[0]
        self.assertEqual(validation.type, "list")
        self.assertEqual(validation.formula1, '"Yes,No"')
        self.assertEqual(validation.prompt, VERIFIED_INPUT_MESSAGE)
        self.assertTrue(validation.showInputMessage)
        self.assertIn("E11", str(validation.sqref))

    def test_multiline_body_prose_has_no_artificial_bullets(self):
        lines = [
            "7. Consider various levels of integration",
            "The religiousness or spirituality of client",
            "The desire of the client to integrate spirituality",
            "The nature of presenting problem",
        ]
        worksheet = self.export_workbook([
            TextParagraph(
                " ".join(lines),
                lines=[text_line(line) for line in lines],
                text_type=TextType.BODY,
            ),
            TextParagraph("One line", text_type=TextType.BODY),
        ])

        self.assertEqual(
            worksheet["B11"].value,
            " ".join(lines),
        )
        self.assertGreater(worksheet.row_dimensions[11].height, worksheet.row_dimensions[12].height)
        self.assertIsNone(worksheet["C11"].value)
        self.assertIsNone(worksheet["D11"].value)
        self.assertIsNone(worksheet["E11"].value)

    def test_genuine_list_markers_remain_intact(self):
        paragraphs = [
            TextParagraph("• first item", text_type=TextType.BULLET),
            TextParagraph("◦ nested item", text_type=TextType.SUB_BULLET),
            TextParagraph("1. numbered item", text_type=TextType.NUMBERED),
        ]
        worksheet = self.export_workbook(paragraphs)

        self.assertEqual(worksheet["B11"].value, "• first item")
        self.assertEqual(worksheet["B12"].value, "◦ nested item")
        self.assertEqual(worksheet["B13"].value, "1. numbered item")

    def test_multiline_title_and_body_paragraphs_have_no_artificial_bullets(self):
        title = TextParagraph(
            "Title continuation prose",
            lines=[text_line("Title"), text_line("continuation prose")],
            text_type=TextType.TITLE,
        )
        body = TextParagraph(
            "Body continuation prose",
            lines=[text_line("Body"), text_line("continuation prose")],
            text_type=TextType.BODY,
        )
        worksheet = self.export_workbook([title, body])

        self.assertEqual(worksheet["B11"].value, title.text)
        self.assertEqual(worksheet["B12"].value, body.text)
        self.assertNotIn("•", worksheet["B11"].value)
        self.assertNotIn("•", worksheet["B12"].value)

    def test_long_text_uses_wrapping_top_alignment_and_a_taller_row(self):
        short_text = "Short text"
        long_text = "Long extracted text " * 35
        worksheet = self.export_workbook([
            TextParagraph(short_text, text_type=TextType.BODY),
            TextParagraph(long_text, text_type=TextType.BODY),
        ])

        self.assertEqual(worksheet["B11"].value, short_text)
        self.assertEqual(worksheet["B12"].value, long_text)
        self.assertTrue(worksheet["B12"].alignment.wrap_text)
        self.assertEqual(worksheet["B12"].alignment.vertical, "top")
        self.assertGreater(
            worksheet.row_dimensions[12].height,
            worksheet.row_dimensions[11].height,
        )

    def test_explicit_newlines_increase_height_and_the_cap_is_applied(self):
        newline_text = "\n".join(["a short source line"] * 8)
        very_long_text = "x" * 5000
        worksheet = self.export_workbook([
            TextParagraph("One line", text_type=TextType.BODY),
            TextParagraph(newline_text, text_type=TextType.BODY),
            TextParagraph(very_long_text, text_type=TextType.BODY),
        ])

        self.assertGreater(
            worksheet.row_dimensions[12].height,
            worksheet.row_dimensions[11].height,
        )
        self.assertEqual(worksheet.row_dimensions[13].height, MAX_CONTENT_ROW_HEIGHT)
        self.assertEqual(
            worksheet["B12"].value,
            newline_text,
        )
        self.assertEqual(worksheet["B13"].value, very_long_text)

    def test_text_columns_are_bounded_and_workbook_reopens_with_unicode(self):
        text = "Résumé\n漢字\nПривет"
        presentation = Presentation(
            metadata={"video_path": "D:/Videos/sample.mp4"},
            slides=[Slide(
                slide_number=1,
                start_time=0.0,
                end_time=1.0,
                paragraphs=[TextParagraph(text, text_type=TextType.BODY)],
            )],
        )

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "export.xlsx"
            export_excel(presentation, str(output_path))
            worksheet = load_workbook(output_path).active

        self.assertLessEqual(worksheet.column_dimensions["B"].width, 80)
        self.assertLessEqual(worksheet.column_dimensions["C"].width, 80)
        self.assertLessEqual(worksheet.column_dimensions["D"].width, 80)
        self.assertLessEqual(worksheet.column_dimensions["E"].width, 20)
        self.assertGreaterEqual(worksheet.column_dimensions["B"].width, 50)
        self.assertEqual(worksheet["B11"].value, text)
        self.assertEqual(worksheet.freeze_panes, "A11")
        self.assertEqual(worksheet.auto_filter.ref, "A10:E11")

    def test_markdown_and_csv_exports_remain_unchanged(self):
        presentation = Presentation(slides=[
            Slide(
                slide_number=1,
                start_time=0.0,
                end_time=1.0,
                paragraphs=[TextParagraph(
                    "Hello continuation",
                    lines=[text_line("Hello"), text_line("continuation")],
                    text_type=TextType.BODY,
                )],
            ),
        ])

        with tempfile.TemporaryDirectory() as temporary_directory:
            output = Path(temporary_directory)
            markdown_path = output / "export.md"
            csv_path = output / "export.csv"
            export_markdown(presentation, str(markdown_path))
            export_csv(presentation, str(csv_path))

            self.assertEqual(
                markdown_path.read_text(encoding="utf-8"),
                "# VideoText Export\n\n# Slide 1\n\nHello continuation\n",
            )
            with csv_path.open(encoding="utf-8", newline="") as file:
                self.assertEqual(list(csv.reader(file)), [
                    ["Slide Number", "Paragraph Type", "Paragraph Text"],
                    ["1", "BODY", "Hello continuation"],
                ])


if __name__ == "__main__":
    unittest.main()
