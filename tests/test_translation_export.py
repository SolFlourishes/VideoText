"""Focused tests for provider-neutral downstream translation exports."""

from __future__ import annotations

import csv
from dataclasses import replace
from pathlib import Path
import sys
import tempfile
import unittest

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from translation_contract import (TranslationProvenance, TranslationRequest, TranslationResult,
    TranslationSourceType, TranslationStatus)
from translation_export import (build_translation_export_records, export_translation_csv,
    export_translation_markdown, export_translation_outputs)
from translation_job import TranslationJob, TranslationOutputGrouping, TranslationOutputPlan, TranslationSourceItem
from translation_output_plan import plan_translation_output
from translation_review import (HumanTranslationReview, HumanTranslationReviewStatus,
    TranslationReviewStatus)
from translation_workbook import TranslationWorkbookRow


def job() -> TranslationJob:
    return TranslationJob("export-job", (
        TranslationSourceItem("video-a", "Video A", "project:a", 0),
        TranslationSourceItem("video-b", "Video B", "project:b", 1),
    ), "en", ("es", "de"), "argos", TranslationOutputPlan(TranslationOutputGrouping.COMBINED))


def evidence(failed: bool = True):
    current = job()
    rows, results = [], []
    for source_index, item in enumerate(current.source_items):
        for language in current.target_languages:
            reference = f"{item.source_item_id}:slide:{source_index + 1}:paragraph:0"
            source = f"Résumé\n漢字 {item.display_name}"
            request = TranslationRequest(f"{reference}:translation:{language}", source, "en", language,
                TranslationProvenance(reference, TranslationSourceType.VERIFIED_OCR), 0)
            rows.append(TranslationWorkbookRow(item.source_item_id, reference, request.request_id, language,
                source_index + 1, source, 0))
            if failed and item.source_item_id == "video-b" and language == "de":
                results.append(TranslationResult(request, TranslationStatus.FAILURE, "argos", error="Network token=hidden"))
            else:
                results.append(TranslationResult(request, TranslationStatus.SUCCESS, "argos", f"{language}: Übersetzung\n完了", model_id="argos-model"))
    return current, tuple(rows), tuple(results)


class TranslationExportTests(unittest.TestCase):
    def test_human_review_preserves_ocr_ai_assessment_and_provenance_layers(self) -> None:
        current, rows, results = evidence(failed=False)
        record = build_translation_export_records(current, rows, results)[0]
        self.assertEqual(HumanTranslationReviewStatus.UNREVIEWED, record.human_review.status)
        self.assertIsNone(record.human_verified_translation)
        self.assertEqual(record.initial_ai_translation, record.reviewed_translation)

        accepted = replace(record, human_review=HumanTranslationReview(
            record.request_id, HumanTranslationReviewStatus.ACCEPTED))
        self.assertEqual(record.initial_ai_translation, accepted.human_verified_translation)
        self.assertEqual(record.initial_ai_translation, accepted.reviewed_translation)

        edited_text = "Traducción revisada por una persona"
        edited = replace(record, human_review=HumanTranslationReview(
            record.request_id, HumanTranslationReviewStatus.EDITED_VERIFIED, edited_text))
        self.assertEqual(edited_text, edited.human_verified_translation)
        self.assertEqual(edited_text, edited.reviewed_translation)
        self.assertEqual(record.original_text, edited.original_text)
        self.assertEqual(record.initial_ai_translation, edited.initial_ai_translation)
        self.assertEqual(record.provider_id, edited.provider_id)
        self.assertEqual(record.model_id, edited.model_id)
        self.assertEqual(record.source_reference, edited.source_reference)
        self.assertEqual(record.review_assessment, edited.review_assessment)

        flagged = replace(record, human_review=HumanTranslationReview(
            record.request_id, HumanTranslationReviewStatus.FLAGGED,
            reviewer_notes="Terminology needs review"))
        self.assertEqual(record.initial_ai_translation, flagged.reviewed_translation)
        self.assertIsNone(flagged.human_verified_translation)
        self.assertEqual(HumanTranslationReviewStatus.FLAGGED, flagged.human_review.status)

    def test_human_review_identity_and_failed_resolution_are_safe(self) -> None:
        current, rows, results = evidence()
        records = build_translation_export_records(current, rows, results)
        with self.assertRaisesRegex(ValueError, "does not match"):
            replace(records[0], human_review=HumanTranslationReview("different-request"))
        failed = replace(records[-1], human_review=HumanTranslationReview(
            records[-1].request_id, HumanTranslationReviewStatus.EDITED_VERIFIED,
            "Human text must not promote a provider failure"))
        self.assertIsNone(failed.reviewed_translation)
        self.assertIsNone(failed.human_verified_translation)
        self.assertEqual(TranslationStatus.FAILURE, failed.translation_status)

    def test_existing_exports_ignore_new_human_review_until_layout_work(self) -> None:
        current, rows, results = evidence(failed=False)
        record = build_translation_export_records(current, rows, results)[0]
        edited = replace(record, human_review=HumanTranslationReview(
            record.request_id, HumanTranslationReviewStatus.EDITED_VERIFIED, "Human revision"))
        with tempfile.TemporaryDirectory() as directory:
            original_path = export_translation_csv((record,), Path(directory) / "original.csv")
            edited_path = export_translation_csv((edited,), Path(directory) / "edited.csv")
            self.assertEqual(
                original_path.read_text(encoding="utf-8"),
                edited_path.read_text(encoding="utf-8"),
            )
            original_markdown = export_translation_markdown(
                (record,), Path(directory) / "original.md")
            edited_markdown = export_translation_markdown(
                (edited,), Path(directory) / "edited.md")
            self.assertEqual(
                original_markdown.read_text(encoding="utf-8"),
                edited_markdown.read_text(encoding="utf-8"),
            )
        self.assertEqual(TranslationReviewStatus.NORMAL_REVIEW, edited.review_assessment.status)

    def test_projection_preserves_exact_evidence_and_failure_has_no_fabricated_translation(self) -> None:
        current, rows, results = evidence()
        records = build_translation_export_records(current, rows, results)
        self.assertEqual(("video-a", "video-a", "video-b", "video-b"), tuple(record.source_item_id for record in records))
        self.assertEqual(("es", "de", "es", "de"), tuple(record.target_language for record in records))
        self.assertEqual("Résumé\n漢字 Video A", records[0].original_text)
        failure = records[-1]
        self.assertIsNone(failure.initial_ai_translation)
        self.assertEqual(TranslationStatus.FAILURE, failure.translation_status)
        self.assertEqual("verified_ocr", failure.source_type.value)
        self.assertNotIn("hidden", failure.error)

    def test_csv_round_trip_has_provenance_unicode_and_blank_failed_translation(self) -> None:
        current, rows, results = evidence()
        records = build_translation_export_records(current, rows, results)
        with tempfile.TemporaryDirectory() as directory:
            path = export_translation_csv(records, Path(directory) / "translations.csv")
            with path.open(encoding="utf-8", newline="") as file:
                values = list(csv.reader(file))
        self.assertEqual(["Source Item", "Slide", "Original Text", "Target Language", "Initial AI Translation", "Translation Status", "Provider", "Model", "Request ID", "Source Reference", "Source Type", "Error", "Review Status", "Review Warnings"], values[0])
        self.assertEqual("Résumé\n漢字 Video A", values[1][2])
        self.assertEqual("es: Übersetzung\n完了", values[1][4])
        self.assertEqual("", values[-1][4]); self.assertEqual("failure", values[-1][5])
        self.assertEqual("verified_ocr", values[-1][10]); self.assertNotIn("hidden", values[-1][11])
        self.assertEqual("Translation Failed", values[-1][12]); self.assertEqual("translation_failed", values[-1][13])

    def test_markdown_represents_success_failure_and_deterministic_sections(self) -> None:
        current, rows, results = evidence()
        records = build_translation_export_records(current, rows, results)
        with tempfile.TemporaryDirectory() as directory:
            path = export_translation_markdown(records, Path(directory) / "translations.md")
            markdown = path.read_text(encoding="utf-8")
        self.assertIn("## Video A — Spanish", markdown)
        self.assertLess(markdown.index("## Video A — Spanish"), markdown.index("## Video B — German"))
        self.assertIn("Résumé\n漢字 Video A", markdown)
        self.assertIn("es: Übersetzung\n完了", markdown)
        self.assertIn("**Translation Status:** Failed", markdown)
        self.assertIn("**Review Status:** Translation Failed", markdown)
        self.assertNotIn("token=hidden", markdown)

    def test_common_export_delegates_excel_and_never_overwrites_text_exports(self) -> None:
        current, rows, results = evidence()
        layout = plan_translation_output(current)
        with tempfile.TemporaryDirectory() as directory:
            summary = export_translation_outputs(current, layout, (row for row in rows), (result for result in results), directory,
                ("csv", "markdown", "excel"))
            self.assertEqual(4, summary.record_count); self.assertEqual(3, summary.success_count); self.assertEqual(1, summary.failure_count)
            self.assertEqual(1, len(summary.paths["excel"]))
            workbook = load_workbook(summary.paths["excel"][0])
            self.assertEqual(["Slide", "Source OCR Text", "Original AI Translation", "Verified Translation",
                              "Human Review Status"],
                [workbook["Video A - Spanish"].cell(10, column).value for column in range(1, 6)])
            self.assertEqual("Translation Review Status", workbook["Video A - Spanish"].cell(10, 6).value)
            with self.assertRaisesRegex(FileExistsError, "already exists"):
                export_translation_outputs(current, layout, rows, results, directory, ("csv",))

    def test_batch_name_is_consistent_across_csv_markdown_and_excel(self) -> None:
        current, rows, results = evidence(failed=False)
        current = replace(
            current,
            output_plan=replace(current.output_plan, batch_name="Mod 2"),
        )
        layout = plan_translation_output(current)
        with tempfile.TemporaryDirectory() as directory:
            summary = export_translation_outputs(
                current, layout, rows, results, directory,
                ("csv", "markdown", "excel"),
            )
            self.assertEqual("Mod 2 - Translation.csv", summary.paths["csv"][0].name)
            self.assertEqual("Mod 2 - Translation.md", summary.paths["markdown"][0].name)
            self.assertEqual("Mod 2 - Translation Batch.xlsx", summary.paths["excel"][0].name)
            with self.assertRaisesRegex(FileExistsError, "already exists"):
                export_translation_outputs(current, layout, rows, results, directory, ("csv",))

    def test_duplicate_missing_and_mismatched_evidence_are_rejected(self) -> None:
        current, rows, results = evidence()
        with self.assertRaisesRegex(ValueError, "Duplicate translation result"):
            build_translation_export_records(current, rows, results + (results[0],))
        with self.assertRaisesRegex(ValueError, "Missing"):
            build_translation_export_records(current, rows, results[:-1])
        invalid = TranslationResult(TranslationRequest(results[0].request_id, "Different", "en", "es",
            results[0].request.provenance, 0), TranslationStatus.SUCCESS, "argos", "Different")
        with self.assertRaisesRegex(ValueError, "does not match"):
            build_translation_export_records(current, rows, (invalid,) + results[1:])
